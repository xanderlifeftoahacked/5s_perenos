import os
import pandas as pd
import openpyxl as ox
import xlwings as xw
import re
import tkinter as tk
from VocabularyApp import VocabularyApp



class DF_parser:
    @classmethod
    def clear_df(cls, df_input: pd.DataFrame):
        cls.df_input = df_input.dropna(how='all')
        cls.df_input = df_input.dropna(axis=1, how='all')

    @classmethod
    def get_grades(cls, df_input: pd.DataFrame):

        cls.grades_set = set()
        cls.grades_list = []

        for column_name in df_input.columns:
            if re.search(r'\d.*[а-яА-Я]', column_name):
                cls.grades_set.add(column_name)
                cls.grades_list.append(column_name)

    @classmethod
    def parse(cls, df_input: pd.DataFrame, df_output: pd.DataFrame, file_in, file_out):
        file_out_temp = 'temporary.xlsx'
        wbxlsm = xw.Book(file_out)
        wbxlsx = ox.Workbook()

        sheet_output_xlsm = wbxlsm.sheets[0]

        wbxlsx.create_sheet('Расстановка')
        sheet_output_xlsx = wbxlsx['Расстановка']
        sheet_output_xlsm.range('A1').number_format = '@'
        cls.clear_df(df_input)
        cls.get_grades(df_input)

        columns = list(cls.df_input.columns)

        last_col_index = max([i for i, col in enumerate(columns) if any(
            c.isalpha() for c in col) and any(c.isdigit() for c in col)])

        cls.df_input = cls.df_input.iloc[:, :last_col_index + 1]

        teachers = [row for index, row in cls.df_input.iterrows()
                    if not pd.isna(row[1])]

        subjects = dict()

        grades_under_five = 0
        for column_name in cls.df_input.columns:
            match = re.search(r'\d+', column_name)
            if match is not None and int(match.group()) <= 4:
                grades_under_five += 1
            
        for teacher in teachers:
            if teacher[1] not in subjects:
                subjects[teacher[1]] = 'nan'

        not_ready = True
        while not_ready:
            root = tk.Tk()
            app = VocabularyApp(root, subjects)
            root.mainloop()
            not_ready = app.is_ready()
        subjects = app.get_vocubalary()

        subject_order = ['Начальные классы','Русский язык и литература', 'Иностранный язык', 
                        'Математика и информатика', 'Общественные науки',
                        'Естественные науки', 'Технология', 'Искусство',
                        'Физическая культура, ОБЖ', 'Курсы по выбору']

        subjects_sorted = sorted(subjects.items(), key=lambda x: subject_order.index(x[1]) if x[1] in subject_order else len(subject_order))
        subjects = dict(subjects_sorted)

        def sort_key(series):
            subject_area = subjects.get(series[1], 'Other')

            for i in range(len(subject_order)):
                if subject_area == subject_order[i]:
                    return i

        teachers = sorted(teachers, key=sort_key)
        # for teacher in teachers:
        #     for subject, area in subjects:
        #         if teacher[1] = subject:

        # print(teachers[0:10])
        # print(cls.df_input.columns)
        # print(cls.grades_set)
        # print(cls.grades_list)
        for grade in range(len(cls.grades_list)):
            sheet_output_xlsx.cell(1, 6 + grade).value = cls.grades_list[grade]

        current_row = 4
        current_col = 2
        teacher_prev_name = 'null'
        subject_area_prev_name = 'null'

        for teacher in teachers:
            if subject_area_prev_name != subjects[teacher[1]]:
                sheet_output_xlsx.cell(row = current_row, column = 1).value = subjects[teacher[1]]
                current_row += 1

            if teacher[0] != teacher_prev_name:
                sheet_output_xlsx.cell(row = current_row, column = current_col).value = teacher[0]
            sheet_output_xlsx.cell(row = current_row, column = current_col + 1).value = teacher[1]

            for item in teacher[2:]:
                current_col += 1
                sheet_output_xlsx.cell(
                    row = current_row, column = current_col + 3).number_format = '@'
                sheet_output_xlsx.cell(row = current_row, column = current_col + 3).value = item

            teacher_prev_name = teacher[0]
            subject_area_prev_name = subjects[teacher[1]]
            current_col = 2
            current_row += 1

        wbxlsx.save(file_out_temp)

        file_in_book = xw.Book(file_out_temp)
        file_in_sheet = file_in_book.sheets[1]

        source_data_range = file_in_sheet.used_range
        to_save_range = sheet_output_xlsm.range("A1:E3").value
        sheet_output_xlsm.range("A1").value = source_data_range.value
        sheet_output_xlsm.range("A1:E3").value = to_save_range

        wbxlsm.save()
        file_in_book.close()
        os.remove(file_out_temp)

# DF_parser.parse(pd.read_excel("C:/Users/User/Downloads/Тарификация_nika ОПШ5.xls"),
#                 pd.read_excel("C:/Users/User/Downloads/upupu/5saplication/Расстановка.xlsm"),
#                 "C:/Users/User/Downloads/upupu/5saplication/Тарификация_nika ОПШ5.xls",
#                 "C:/Users/User/Downloads/upupu/5saplication/Расстановка.xlsm")

# workbook_xlsx = xw.Book('temporary.xlsx')
# sheet_xlsx = workbook_xlsx.sheets[1]

# # Open the destination workbook and sheet (in this example, same workbook)
# workbook_xlsm = xw.Book('Расстановка.xlsm')
# sheet_xlsm = workbook_xlsm.sheets[0]

# # Get the range of data from the source sheet
# source_data_range = sheet_xlsx.used_range

# # Paste the data into the destination sheet
# sheet_xlsm.range('A1').value = source_data_range.value

# # Save the changes made to the workbook
# workbook_xlsm.save()